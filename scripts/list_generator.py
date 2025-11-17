#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
加分名单生成脚本
功能：按模板生成标准加分名单Excel，严格遵循格式规范
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import argparse
import os
import sys

# 模板配置
TEMPLATES = {
    'activity': {
        'columns': [
            {'label': '序号', 'field': '序号', 'width': 4.75},
            {'label': '班级', 'field': '行政班级', 'width': 27},
            {'label': '姓名', 'field': '姓名', 'width': 8.25}
        ],
        'rowHeights': {
            'title': 69,
            'note': 24.5,
            'header': 20,
            'body': 20
        }
    },
    'competition': {
        'columns': [
            {'label': '序号', 'field': '序号', 'width': 9.25},
            {'label': '班级', 'field': '行政班级', 'width': 30},
            {'label': '姓名', 'field': '姓名', 'width': 17.25},
            {'label': '奖项', 'field': '奖项', 'width': 16.25},
            {'label': '加分数', 'field': '加分分数', 'width': 7.25}
        ],
        'rowHeights': {
            'title': 69,
            'note': 24.5,
            'header': 20,
            'body': 20
        }
    }
}

def load_data(input_file):
    """加载筛选后的数据"""
    try:
        df = pd.read_excel(input_file, sheet_name='筛选结果')
        return df
    except Exception as e:
        print(f"加载数据失败: {str(e)}")
        sys.exit(1)

def generate_list(data, template_type, semester, activity_name, score_type, score=None):
    """生成加分名单"""
    template = TEMPLATES[template_type]
    
    # 创建Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '加分名单'
    
    # 设置字体
    title_font = Font(name='微软雅黑', size=18, bold=True)
    note_font = Font(name='微软雅黑', size=12)
    header_font = Font(name='微软雅黑', size=11, bold=True)
    body_font = Font(name='微软雅黑', size=11)
    
    # 设置对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 1
    
    # 生成标题
    if template_type == 'competition':
        title = f"{semester}工学院{activity_name}加分名单"
    else:
        title = f"{semester}{activity_name}加分名单"
    
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(template["columns"]))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=title)
    cell.font = title_font
    cell.alignment = center_alignment
    ws.row_dimensions[current_row].height = template['rowHeights']['title']
    current_row += 1
    
    # 生成注释
    if template_type == 'competition':
        note = f"注：以下同学加{score_type}，具体分数如下"
    else:
        note = f"注：以下同学每人加{score_type}{score}分"
    
    ws.merge_cells(f'A{current_row}:{get_column_letter(len(template["columns"]))}{current_row}')
    cell = ws.cell(row=current_row, column=1, value=note)
    cell.font = note_font
    cell.alignment = center_alignment
    ws.row_dimensions[current_row].height = template['rowHeights']['note']
    current_row += 1
    
    # 生成表头
    headers = [col['label'] for col in template['columns']]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
    ws.row_dimensions[current_row].height = template['rowHeights']['header']
    current_row += 1
    
    # 生成数据行
    for idx, row_data in enumerate(data.itertuples(), start=1):
        for col_idx, col_config in enumerate(template['columns'], start=1):
            if col_config['field'] == '序号':
                value = idx
            else:
                value = getattr(row_data, col_config['field'], '')
            
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = center_alignment
        
        ws.row_dimensions[current_row].height = template['rowHeights']['body']
        current_row += 1
    
    # 设置列宽
    for col_idx, col_config in enumerate(template['columns'], start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_config['width']
    
    return wb

def main():
    parser = argparse.ArgumentParser(description='加分名单生成脚本')
    parser.add_argument('--input', type=str, required=True, help='输入文件路径（filtered_data.xlsx）')
    parser.add_argument('--template', type=str, required=True, choices=['activity', 'competition'], help='模板类型：activity（活动/讲座）或competition（比赛）')
    parser.add_argument('--semester', type=str, required=True, help='学年学期，如：2024-2025学年第X学期')
    parser.add_argument('--activity', type=str, required=True, help='活动名称')
    parser.add_argument('--scoreType', type=str, required=True, help='加分类型：学业分或品德分')
    parser.add_argument('--score', type=float, help='加分分数（仅活动/讲座类需要）')
    parser.add_argument('--output', type=str, default='./output', help='输出目录')
    
    args = parser.parse_args()
    
    # 创建输出目录
    os.makedirs(args.output, exist_ok=True)
    
    # 加载数据
    df = load_data(args.input)
    print(f"加载数据: {len(df)} 行")
    
    # 验证参数
    if args.template == 'activity' and args.score is None:
        print("错误: 活动/讲座类模板需要指定--score参数")
        sys.exit(1)
    
    # 生成文件名
    filename = f"{args.semester}工学院{args.activity}加分名单.xlsx"
    output_path = os.path.join(args.output, filename)
    
    # 生成名单
    wb = generate_list(
        df,
        args.template,
        args.semester,
        args.activity,
        args.scoreType,
        args.score
    )
    
    # 保存文件
    wb.save(output_path)
    print(f"已生成加分名单到: {output_path}")
    print("\n生成完成！")

if __name__ == '__main__':
    main()

